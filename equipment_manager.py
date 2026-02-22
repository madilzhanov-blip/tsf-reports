# -*- coding: utf-8 -*-
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

class EquipmentManager:
    """Менеджер для управления данными по технике"""
    
    # Типы техники с их характеристиками
    EQUIPMENT_TYPES = {
        "Каток": {"max_count": 74, "unit": "шт", "category": "Дорожная техника"},
        "Самосвал 25": {"max_count": 62, "unit": "шт", "category": "Транспорт"},
        "Самосвал 75": {"max_count": 78, "unit": "шт", "category": "Транспорт"},
        "Самосвал 150": {"max_count": 10, "unit": "шт", "category": "Транспорт"},
        "Бульдозер": {"max_count": 56, "unit": "шт", "category": "Земляные работы"},
        "Экскаватор": {"max_count": 65, "unit": "шт", "category": "Земляные работы"},
        "Грейдер": {"max_count": 7, "unit": "шт", "category": "Дорожная техника"},
        "Погрузчик": {"max_count": 9, "unit": "шт", "category": "Погрузочная техника"},
        "Водовоз": {"max_count": 32, "unit": "шт", "category": "Специальная техника"}
    }
    
    # Статусы техники
    EQUIPMENT_STATUSES = [
        "В работе",
        "В ожидании",
        "В ремонте", 
        "Без механизатора",
        "Нет топлива",
        "Простой"
    ]
    
    # Участки работ
    WORK_LOCATIONS = [
        "На Дамбе",
        "Карьер Шархия", 
        "Магистральный пульповод",
        "Водовод оборотной воды",
        "Межплощадочные сети ВЛ 110",
        "ГПП-1 ПС 110/10кВ",
        "Распределительный пульповод",
        "Хвостохранилище"
    ]
    
    def __init__(self, data_file="equipment_reports.json"):
        self.data_file = data_file
        self.reports = self.load_reports()
    
    def load_reports(self):
        """Загрузка отчетов из файла"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def save_reports(self):
        """Сохранение отчетов в файл"""
        try:
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(self.reports, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"Ошибка сохранения: {e}")
            return False
    
    def create_shift_report(self, date, shift, equipment_data, inspector_name):
        """Создание сменного отчета по технике"""
        report = {
            "id": len(self.reports) + 1,
            "date": date,
            "shift": shift,  # "День" или "Ночь"
            "inspector_name": inspector_name,
            "equipment_data": equipment_data,
            "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "totals": self.calculate_totals(equipment_data)
        }
        
        self.reports.append(report)
        self.save_reports()
        return report
    
    def calculate_totals(self, equipment_data):
        """Расчет итоговых показателей"""
        totals = {
            "total_available": 0,
            "total_working": 0,
            "total_waiting": 0,
            "total_repair": 0,
            "total_no_operator": 0,
            "total_no_fuel": 0,
            "efficiency_percent": 0
        }
        
        for equipment in equipment_data:
            totals["total_available"] += equipment.get("available", 0)
            totals["total_working"] += equipment.get("working", 0)
            totals["total_waiting"] += equipment.get("waiting", 0)
            totals["total_repair"] += equipment.get("repair", 0)
            totals["total_no_operator"] += equipment.get("no_operator", 0)
            totals["total_no_fuel"] += equipment.get("no_fuel", 0)
        
        # Расчет эффективности
        if totals["total_available"] > 0:
            totals["efficiency_percent"] = round(
                (totals["total_working"] / totals["total_available"]) * 100, 1
            )
        
        return totals
    
    def get_reports_by_period(self, date_from=None, date_to=None, shift=None):
        """Получение отчетов за период"""
        filtered_reports = self.reports
        
        if date_from:
            filtered_reports = [r for r in filtered_reports if r["date"] >= date_from]
        
        if date_to:
            filtered_reports = [r for r in filtered_reports if r["date"] <= date_to]
            
        if shift:
            filtered_reports = [r for r in filtered_reports if r["shift"] == shift]
        
        return sorted(filtered_reports, key=lambda x: (x["date"], x["shift"]), reverse=True)
    
    def get_equipment_statistics(self, days=30):
        """Статистика по технике за период"""
        end_date = datetime.datetime.now().strftime("%Y-%m-%d")
        start_date = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
        
        reports = self.get_reports_by_period(start_date, end_date)
        
        stats = {}
        for equipment_type in self.EQUIPMENT_TYPES.keys():
            stats[equipment_type] = {
                "total_shifts": 0,
                "avg_working": 0,
                "avg_efficiency": 0,
                "total_working_hours": 0
            }
        
        # Подсчет статистики
        for report in reports:
            for equipment in report["equipment_data"]:
                eq_type = equipment["type"]
                if eq_type in stats:
                    stats[eq_type]["total_shifts"] += 1
                    stats[eq_type]["total_working_hours"] += equipment.get("working", 0)
        
        # Расчет средних значений
        for eq_type in stats:
            if stats[eq_type]["total_shifts"] > 0:
                stats[eq_type]["avg_working"] = round(
                    stats[eq_type]["total_working_hours"] / stats[eq_type]["total_shifts"], 1
                )
                
                max_available = self.EQUIPMENT_TYPES[eq_type]["max_count"]
                if max_available > 0:
                    stats[eq_type]["avg_efficiency"] = round(
                        (stats[eq_type]["avg_working"] / max_available) * 100, 1
                    )
        
        return stats
    
    def create_excel_report(self, date_from, date_to):
        """Создание Excel отчета за период"""
        reports = self.get_reports_by_period(date_from, date_to)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Техника {date_from} - {date_to}"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Заголовки
        headers = [
            "Дата", "Смена", "Техника", "Общее кол-во", "В работе", 
            "% загрузки", "В ожидании", "В ремонте", "Без механизатора", 
            "Нет топлива", "На дамбе", "Участок работ", "Инспектор"
        ]
        
        # Заполнение заголовков
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Заполнение данных
        row = 2
        for report in reports:
            for equipment in report["equipment_data"]:
                ws.cell(row=row, column=1, value=report["date"])
                ws.cell(row=row, column=2, value=report["shift"])
                ws.cell(row=row, column=3, value=equipment["type"])
                ws.cell(row=row, column=4, value=equipment.get("available", 0))
                ws.cell(row=row, column=5, value=equipment.get("working", 0))
                
                # Расчет процента загрузки
                available = equipment.get("available", 0)
                working = equipment.get("working", 0)
                efficiency = round((working / available * 100), 1) if available > 0 else 0
                ws.cell(row=row, column=6, value=efficiency)
                
                ws.cell(row=row, column=7, value=equipment.get("waiting", 0))
                ws.cell(row=row, column=8, value=equipment.get("repair", 0))
                ws.cell(row=row, column=9, value=equipment.get("no_operator", 0))
                ws.cell(row=row, column=10, value=equipment.get("no_fuel", 0))
                ws.cell(row=row, column=11, value=equipment.get("on_dam", 0))
                ws.cell(row=row, column=12, value=equipment.get("location", ""))
                ws.cell(row=row, column=13, value=report["inspector_name"])
                
                # Применяем границы
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).border = border
                
                row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def get_dashboard_data(self):
        """Данные для дашборда"""
        # Последний отчет
        latest_report = self.reports[-1] if self.reports else None
        
        # Статистика за 30 дней
        stats = self.get_equipment_statistics(30)
        
        # Тренды по дням (последние 7 дней)
        trends = {}
        for i in range(7):
            date = (datetime.datetime.now() - datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            day_reports = self.get_reports_by_period(date, date)
            
            day_efficiency = 0
            day_working = 0
            day_available = 0
            
            for report in day_reports:
                day_working += report["totals"]["total_working"]
                day_available += report["totals"]["total_available"]
            
            if day_available > 0:
                day_efficiency = round((day_working / day_available) * 100, 1)
            
            trends[date] = {
                "efficiency": day_efficiency,
                "working": day_working,
                "available": day_available
            }
        
        return {
            "latest_report": latest_report,
            "statistics": stats,
            "trends": trends,
            "total_reports": len(self.reports)
        }

def test_equipment_manager():
    """Тестирование модуля техники"""
    print("🧪 Тестирование модуля учета техники...")
    
    # Создаем менеджер
    eq_manager = EquipmentManager()
    
    # Создаем тестовые данные
    test_equipment_data = [
        {
            "type": "Каток",
            "available": 70,
            "working": 43,
            "waiting": 5,
            "repair": 2,
            "no_operator": 15,
            "no_fuel": 5,
            "on_dam": 43,
            "location": "На Дамбе"
        },
        {
            "type": "Самосвал 25",
            "available": 40,
            "working": 12,
            "waiting": 8,
            "repair": 5,
            "no_operator": 10,
            "no_fuel": 5,
            "on_dam": 12,
            "location": "Карьер Шархия"
        },
        {
            "type": "Экскаватор",
            "available": 44,
            "working": 34,
            "waiting": 3,
            "repair": 2,
            "no_operator": 3,
            "no_fuel": 2,
            "on_dam": 34,
            "location": "На Дамбе"
        }
    ]
    
    # Создаем тестовый отчет
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    report = eq_manager.create_shift_report(
        date=today,
        shift="Ночь",
        equipment_data=test_equipment_data,
        inspector_name="Мадияр Адильжанов"
    )
    
    print(f"✅ Создан отчет #{report['id']}")
    print(f"📊 Общая эффективность: {report['totals']['efficiency_percent']}%")
    print(f"🚜 Всего в работе: {report['totals']['total_working']}")
    print(f"⏳ В ожидании: {report['totals']['total_waiting']}")
    
    # Тестируем статистику
    stats = eq_manager.get_equipment_statistics(7)
    print(f"\n📈 Статистика за 7 дней:")
    for eq_type, data in stats.items():
        if data["total_shifts"] > 0:
            print(f"  {eq_type}: {data['avg_efficiency']}% эффективность")
    
    # Тестируем дашборд
    dashboard = eq_manager.get_dashboard_data()
    print(f"\n📋 Всего отчетов: {dashboard['total_reports']}")
    
    print("🎉 Тестирование завершено!")

if __name__ == "__main__":
    test_equipment_manager()