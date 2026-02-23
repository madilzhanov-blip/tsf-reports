from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file
import io
from flask import Flask, render_template, request, redirect, url_for, flash, session
import json
import os
from functools import wraps
import requests
from werkzeug.utils import secure_filename
from datetime import datetime



#для фото NCR
UPLOAD_FOLDER = 'static/uploads/ncr_photos'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Конфигурация для погоды
WEATHER_CONFIG = {
    'api_key': '9d466af824eaf8a62e14e003bc825247',
    'lat': 40.8606,  # Алмалык
    'lon': 69.6047,
    'cache_duration': 3600  # 1 час
}


import time

weather_cache = {'data': None, 'timestamp': 0}

def get_weather_data():
    current_time = time.time()

    if (
        weather_cache['data']
        and current_time - weather_cache['timestamp'] < WEATHER_CONFIG['cache_duration']
    ):
        return weather_cache['data']

    try:
        url = "https://api.openweathermap.org/data/2.5/weather"
        params = {
            'lat': WEATHER_CONFIG['lat'],
            'lon': WEATHER_CONFIG['lon'],
            'appid': WEATHER_CONFIG['api_key'],
            'units': 'metric',
            'lang': 'ru'
        }

        response = requests.get(url, params=params, timeout=5)
        data = response.json()

        weather_text = (
            f"{data['weather'][0]['description'].title()}, "
            f"{round(data['main']['temp'])}°C, "
            f"влажность {data['main']['humidity']}%, "
            f"ветер {data['wind']['speed']} м/с"
        )

        weather_cache['data'] = weather_text
        weather_cache['timestamp'] = current_time

        return weather_text

    except Exception as e:
        print(f"Ошибка получения погоды: {e}")
        return "Не удалось получить данные о погоде"

try:
    from word_export_manager import DailyReportWordExporter
    WORD_EXPORT_AVAILABLE = True
    print("✅ Word Export модуль загружен")
except ImportError as e:
    WORD_EXPORT_AVAILABLE = False
    print(f"⚠️ Word Export модуль не найден: {e}")
except Exception as e:
    WORD_EXPORT_AVAILABLE = False
    print(f"⚠️ Ошибка загрузки Word Export: {e}")




try:
    from file_export_manager import FileExportManager, EmailSender, EMAIL_CONFIGS
    
    # Настройки экспорта
    EXPORT_CONFIG = {
        'local_folder': 'exports',
        'network_folder': None, 
        'email_enabled': False,  # Пока отключено
        'email_config': 'corporate'  # gmail, outlook, corporate
    }
    
    
    export_manager = FileExportManager(
        export_folder='exports'
    )
    
    print("✅ Система экспорта загружена")
    
except ImportError as e:
    export_manager = None
    print(f"⚠️ Система экспорта недоступна: {e}")




app = Flask(__name__)
app.secret_key = 'inspection_system_2025'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

sp_connector = None

if all([
    os.getenv("SP_TENANT_ID"),
    os.getenv("SP_CLIENT_ID"),
    os.getenv("SP_CLIENT_SECRET"),
    os.getenv("SP_SITE_URL"),
]):
    try:
        from sharepoint_connector import SharePointConnector
        sp_connector = SharePointConnector(
            tenant_id=os.getenv("SP_TENANT_ID"),
            client_id=os.getenv("SP_CLIENT_ID"),
            client_secret=os.getenv("SP_CLIENT_SECRET"),
            site_url=os.getenv("SP_SITE_URL"),
        )
        print("✅ SharePoint подключен")
    except Exception as e:
        sp_connector = None
        print("⚠️ SharePoint отключен:", e)
        
DATA_FILES = {
    'users': 'inspectors.json',
    'geodetic_inspections': 'geodetic_inspections.json',
    'civil_inspections': 'civil_inspections.json',
    '_reports': '_reports.json',
    'remark_inspections': 'remark_inspections.json',
    'daily_reports': 'daily_reports.json' 
}

class DataManager:
    @staticmethod
    def load_data(data_type):
        filename = DATA_FILES.get(data_type, f'{data_type}.json')
        if os.path.exists(filename):
            try:
                with open(filename, 'r', encoding='utf-8') as file:
                    return json.load(file)
            except:
                return []
        return []
    
    @staticmethod
    def save_data(data_type, data):
        filename = DATA_FILES.get(data_type, f'{data_type}.json')
        try:
            with open(filename, 'w', encoding='utf-8') as file:
                json.dump(data, file, ensure_ascii=False, indent=2)
            return True
        except:
            return False
    
    @staticmethod
    def add_record(data_type, record):
        data = DataManager.load_data(data_type)
        record['id'] = max([r['id'] for r in data], default=0) + 1
        record['created_at'] = datetime.now().isoformat()
        data.append(record)
        return DataManager.save_data(data_type, data)

    @staticmethod
    def find_user(username, password):
        users = DataManager.load_data('users')
        for user in users:
            if user['username'] == username and user['password'] == password:
                return user
        return None
    
    @staticmethod
    def update_record(data_type, record_id, updated_data):
        """Обновление записи по ID"""
        data = DataManager.load_data(data_type)
        for i, record in enumerate(data):
            if record['id'] == record_id:
                
                updated_data['id'] = record['id']
                updated_data['created_at'] = record['created_at']
                updated_data['updated_at'] = datetime.now().isoformat()
                data[i] = updated_data
                return DataManager.save_data(data_type, data)
        return False
    
    @staticmethod
    def get_record_by_id(data_type, record_id):
        """Получение записи по ID"""
        data = DataManager.load_data(data_type)
        for record in data:
            if record['id'] == record_id:
                return record
        return None
    
    @staticmethod
    def delete_record(data_type, record_id):
        """Удаление записи по ID"""
        data = DataManager.load_data(data_type)
        original_length = len(data)
        
        data = [record for record in data if record['id'] != record_id]
    
        if len(data) < original_length:
            return DataManager.save_data(data_type, data)
        return False  # Запись не найдена
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('⚠️ Необходимо войти в систему', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function






@app.route('/export_inspection/<inspection_type>/<int:inspection_id>')
@login_required
def export_single_inspection(inspection_type, inspection_id):
    """Экспорт одной инспекции в файл"""
    if not export_manager:
        flash('❌ Система экспорта недоступна', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        
        data_key = f'{inspection_type}_inspections'
        inspection = DataManager.get_record_by_id(data_key, inspection_id)
        
        if not inspection:
            flash('❌ Инспекция не найдена', 'error')
            return redirect(url_for('dashboard'))
        
        
        wb = create_excel_file([inspection], inspection_type.upper())
        
       
        success, filepath, filename = export_manager.save_excel_file(
            wb, inspection_type, inspection_id
        )
        
        if success:
         
            if EXPORT_CONFIG.get('network_folder'):
                net_success, net_message = export_manager.copy_to_network_folder(filepath, filename)
                if net_success:
                    flash(f'✅ Файл сохранен локально и в сети: {filename}', 'success')
                else:
                    flash(f'✅ Файл сохранен локально: {filename}', 'success')
            else:
                flash(f'✅ Файл сохранен: {filename}', 'success')
        else:
            flash('❌ Ошибка сохранения файла', 'error')
            
    except Exception as e:
        flash(f'❌ Ошибка экспорта: {str(e)}', 'error')
    
    return redirect(url_for(f'{inspection_type}_list'))
    
@app.route('/test_ncr_number')
def test_ncr_number():
    return generate_ncr_number()
    
@app.route('/test_export_system')
@login_required  
def test_export_system():
    """Тестирование системы экспорта"""
    if not export_manager:
        flash('❌ Система экспорта недоступна', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        from file_export_manager import test_export_system
        test_export_system()
        flash('✅ Система экспорта работает! Проверь папку exports/', 'success')
    except Exception as e:
        flash(f'❌ Ошибка тестирования: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))



def create_excel_file(data, inspection_type):
    """Создание Excel файла с данными инспекций"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{inspection_type} инспекции"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
   
    if inspection_type == "Геодезические":
        headers = [
            "№", "Проект", "Укрупненный объект", "Пакет работ", "Дата инспекции",
            "Объект", "Участок", "Название работ", "Кол-во по проекту", 
            "Кол-во фактически", "Пикетаж от", "Пикетаж до", "ФИО Инспектора",
            "Решение", "Комментарии", "Кто создал", "Дата создания"
        ]
        
       
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        
        for row, inspection in enumerate(data, 2):
            ws.cell(row=row, column=1, value=inspection.get('id', ''))
            ws.cell(row=row, column=2, value=inspection.get('project', ''))
            ws.cell(row=row, column=3, value=inspection.get('major_object', ''))
            ws.cell(row=row, column=4, value=inspection.get('work_package', ''))
            ws.cell(row=row, column=5, value=inspection.get('inspection_date', ''))
            ws.cell(row=row, column=6, value=inspection.get('object', ''))
            ws.cell(row=row, column=7, value=inspection.get('section', ''))
            ws.cell(row=row, column=8, value=inspection.get('work_name', ''))
            ws.cell(row=row, column=9, value=inspection.get('quantity_project', ''))
            ws.cell(row=row, column=10, value=inspection.get('quantity_actual', ''))
            ws.cell(row=row, column=11, value=inspection.get('picket_from', ''))
            ws.cell(row=row, column=12, value=inspection.get('picket_to', ''))
            ws.cell(row=row, column=13, value=inspection.get('inspector_name_field', ''))
            ws.cell(row=row, column=14, value=inspection.get('decision', ''))
            ws.cell(row=row, column=15, value=inspection.get('deviation_comments', ''))
            ws.cell(row=row, column=16, value=inspection.get('inspector_name', ''))
            ws.cell(row=row, column=17, value=inspection.get('created_at', ''))
    
    
    elif inspection_type == "":
        headers = [
            "№", "Номер ", "Дисциплина", "Наименование объекта", "Конструктив / расположение / участок", 
            "Описание несоответствия", "Дата выпуска", "Корректирующие действия Подрядчика", "Планируемая дата закрытия", 
            "Фактическая дата закрытия",  "ФИО Инспектора", "Статус ", "Кто создал", "Дата создания"
        ]
        
   
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        

        for row,  in enumerate(data, 2):
            ws.cell(row=row, column=1, value=.get('id', ''))
            ws.cell(row=row, column=2, value=.get('_Number', ''))
            ws.cell(row=row, column=3, value=.get('Discipline', ''))
            ws.cell(row=row, column=4, value=.get('major_object', ''))
            ws.cell(row=row, column=5, value=.get('Location', ''))
            ws.cell(row=row, column=6, value=.get('_Description', ''))
            ws.cell(row=row, column=7, value=.get('inspection_date', ''))
            ws.cell(row=row, column=8, value=.get('Correction_acts', ''))
            ws.cell(row=row, column=9, value=.get('closed_date.plan', ''))
            ws.cell(row=row, column=10, value=.get('closed_date.actual', ''))
            ws.cell(row=row, column=11, value=.get('inspector_name_field', ''))
            ws.cell(row=row, column=12, value=.get('_Status', ''))
            ws.cell(row=row, column=13, value=.get('inspector_name', ''))
            ws.cell(row=row, column=14, value=.get('created_at', ''))
    
    
    elif inspection_type == "REMARK":
        headers = [
            "№", "Дисциплина", "Наименование объекта", "Описание замечания", 
            "Ссылка на нормативный документ", "Дата выпуска", "Ответственное лицо от подрядчика", 
            "ФИО Инспектора", "Планируемая дата закрытия", "Фактическая дата закрытия", 
            "Причина отклонения и действия по устранению", "Статус", "Кто создал", "Дата создания"
        ]
        
      
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
      
        for row, remark in enumerate(data, 2):
            ws.cell(row=row, column=1, value=remark.get('id', ''))
            ws.cell(row=row, column=2, value=remark.get('Discipline', ''))
            ws.cell(row=row, column=3, value=remark.get('major_object', ''))
            ws.cell(row=row, column=4, value=remark.get('remark_description', ''))
            ws.cell(row=row, column=5, value=remark.get('link_to_normative', ''))
            ws.cell(row=row, column=6, value=remark.get('inspection_date', ''))
            ws.cell(row=row, column=7, value=remark.get('responsible_person_cont', ''))
            ws.cell(row=row, column=8, value=remark.get('inspector_name_field', ''))
            ws.cell(row=row, column=9, value=remark.get('remark_closed_date.plan', ''))
            ws.cell(row=row, column=10, value=remark.get('remark_closed_date.actual', ''))
            ws.cell(row=row, column=11, value=remark.get('reason_for_reject', ''))
            ws.cell(row=row, column=12, value=remark.get('Status', ''))
            ws.cell(row=row, column=13, value=remark.get('inspector_name', ''))
            ws.cell(row=row, column=14, value=remark.get('created_at', ''))
    

    elif inspection_type == "CIVIL":
        headers = [
            "№", "Проект", "Укрупненный объект", "Пакет работ", "Объект",
            "Участок", "Название работы", "Подробное описание", "№ Лаб. протокола",
            "Количество", "Маркировка", "Номер RFI", "Пикетаж от", "Пикетаж до",
            "Статус RFI", "Дата инспекции", "ФИО Инспектора", "Решение", "Комментарии",
            "№ ИТД", "Кто создал", "Дата создания"                          
        ]
        
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        
        for row, inspection in enumerate(data, 2):
            ws.cell(row=row, column=1, value=inspection.get('id', ''))
            ws.cell(row=row, column=2, value=inspection.get('project', ''))
            ws.cell(row=row, column=3, value=inspection.get('major_object', ''))
            ws.cell(row=row, column=4, value=inspection.get('work_package', ''))
            ws.cell(row=row, column=5, value=inspection.get('object', ''))
            ws.cell(row=row, column=6, value=inspection.get('section', ''))
            ws.cell(row=row, column=7, value=inspection.get('work_name', ''))
            ws.cell(row=row, column=8, value=inspection.get('work_description', ''))
            ws.cell(row=row, column=9, value=inspection.get('laboratory_number', ''))
            ws.cell(row=row, column=10, value=inspection.get('quantity', ''))
            ws.cell(row=row, column=11, value=inspection.get('marking', ''))
            ws.cell(row=row, column=12, value=inspection.get('RFI_number', ''))
            ws.cell(row=row, column=13, value=inspection.get('picket_from', ''))
            ws.cell(row=row, column=14, value=inspection.get('picket_to', ''))
            ws.cell(row=row, column=15, value=inspection.get('RFI_status', ''))
            ws.cell(row=row, column=16, value=inspection.get('inspection_date', ''))
            ws.cell(row=row, column=17, value=inspection.get('inspector_name_field', ''))
            ws.cell(row=row, column=18, value=inspection.get('decision', ''))
            ws.cell(row=row, column=19, value=inspection.get('deviation_comments', ''))
            ws.cell(row=row, column=20, value=inspection.get('executive_number', ''))
            ws.cell(row=row, column=21, value=inspection.get('inspector_name', ''))
            ws.cell(row=row, column=22, value=inspection.get('created_at', ''))
    
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


@app.route('/export_remark')
@login_required
def export_remark():
    """Экспорт замечаний в Excel"""
    remark_reports = DataManager.load_data('remark_inspections')
    
    
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        remark_reports = [r for r in remark_reports if r.get('Status') == status_filter]
    if date_from:
        remark_reports = [r for r in remark_reports if r.get('inspection_date', '') >= date_from]
    if date_to:
        remark_reports = [r for r in remark_reports if r.get('inspection_date', '') <= date_to]
    
    wb = create_excel_file(remark_reports, "REMARK")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"remark_reports_{datetime.now().isoformat()}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export_geodetic')
@login_required
def export_geodetic():
    """Экспорт геодезических инспекций в Excel"""
    inspections = DataManager.load_data('geodetic_inspections')
    
    
    status_filter = request.args.get('status', '')
    decision_filter = request.args.get('decision', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        inspections = [i for i in inspections if i.get('status') == status_filter]
    if decision_filter:
        inspections = [i for i in inspections if i.get('decision') == decision_filter]
    if date_from:
        inspections = [i for i in inspections if i.get('inspection_date', '') >= date_from]
    if date_to:
        inspections = [i for i in inspections if i.get('inspection_date', '') <= date_to]
    
    wb = create_excel_file(inspections, "Геодезические")
    

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"geodetic_inspections_{datetime.now().isoformat()}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export_')
@login_required
def export_ncr():
    """Экспорт NCR в Excel"""
    ncr_reports = DataManager.load_data('ncr_reports')
    
    
    status_filter = request.args.get('NCR_status', '')
    priority_filter = request.args.get('priority', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        ncr_reports = [r for r in ncr_reports if r.get('NCR_status') == status_filter]
    if priority_filter:
        ncr_reports = [r for r in ncr_reports if r.get('priority') == priority_filter]
    if date_from:
        ncr_reports = [r for r in ncr_reports if r.get('inspection_date', '') >= date_from]
    if date_to:
        ncr_reports = [r for r in ncr_reports if r.get('inspection_date', '') <= date_to]
    
    wb = create_excel_file(ncr_reports, "NCR")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ncr_reports_{datetime.now().isoformat()}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export_civil')
@login_required
def export_civil():
    """Экспорт монтажных инспекций в Excel"""
    inspections = DataManager.load_data('civil_inspections')
    
    
    status_filter = request.args.get('status', '')
    decision_filter = request.args.get('decision', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        inspections = [i for i in inspections if i.get('status') == status_filter]
    if decision_filter:
        inspections = [i for i in inspections if i.get('decision') == decision_filter]
    if date_from:
        inspections = [i for i in inspections if i.get('inspection_date', '') >= date_from]
    if date_to:
        inspections = [i for i in inspections if i.get('inspection_date', '') <= date_to]
    
    wb = create_excel_file(inspections, "CIVIL")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"civil_inspections_{datetime.now().isoformat()}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



@app.route('/edit_geodetic/<int:inspection_id>', methods=['GET', 'POST'])
@login_required
def edit_geodetic(inspection_id):
    """Редактирование геодезической инспекции"""
    inspection = DataManager.get_record_by_id('geodetic_inspections', inspection_id)
    
    if not inspection:
        flash('❌ Инспекция не найдена', 'error')
        return redirect(url_for('geodetic_list'))
    
    if request.method == 'POST':
        updated_inspection = {
            'project': request.form.get('project', ''),
            'major_object': request.form.get('major_object', ''),
            'work_package': request.form.get('work_package', ''),
            'inspection_date': request.form.get('inspection_date', ''),
            'object': request.form.get('object', ''),
            'section': request.form.get('section', ''),
            'work_name': request.form.get('work_name', ''),
            'quantity_project': request.form.get('quantity_project', ''),
            'quantity_actual': request.form.get('quantity_actual', ''),
            'picket_from': request.form.get('picket_from', ''),
            'picket_to': request.form.get('picket_to', ''),
            'inspector_name_field': request.form.get('inspector_name_field', ''),
            'decision': request.form.get('decision', ''),
            'deviation_comments': request.form.get('deviation_comments', ''),
            'inspector_id': inspection['inspector_id'],  
            'inspector_name': inspection['inspector_name']
        }
        
        if DataManager.update_record('geodetic_inspections', inspection_id, updated_inspection):
            flash('✅ Геодезическая инспекция обновлена!', 'success')
            return redirect(url_for('geodetic_list'))
        else:
            flash('❌ Ошибка при обновлении', 'error')
    
    return render_template('edit_geodetic.html', inspection=inspection)



@app.route('/edit_ncr/<int:ncr_id>', methods=['GET', 'POST'])
@login_required
def edit_ncr(ncr_id):
    ncr = DataManager.get_record_by_id('ncr_reports', ncr_id)

    if not ncr:
        flash('❌ Отчет не найден', 'error')
        return redirect(url_for('ncr_list'))

    existing_photos = ncr.get('photos', [])
    

    if request.method == 'POST':

        # ─── УДАЛЕНИЕ ССЫЛКИ НА ПОДПИСАННЫЙ СКАН ───
        if request.form.get('remove_signed_scan') == '1':
            ncr['signed_scan_url'] = None
            ncr['signed_scan_uploaded_at'] = None
            ncr['signed_scan_uploaded_by'] = None

            DataManager.update_record('ncr_reports', ncr_id, ncr)
            flash('🗑️ Ссылка на подписанный скан удалена', 'success')
            return redirect(url_for('edit_ncr', ncr_id=ncr_id))

                # ─── ЗАКРЫТИЕ NCR ───
        if request.form.get('close_ncr') == '1':

            if not ncr.get('signed_scan_url'):
                flash('❌ Нельзя закрыть NCR без подписанного скана', 'error')
                return redirect(url_for('edit_ncr', ncr_id=ncr_id))

            ncr['NCR_Status'] = 'Закрыто'
            ncr['closed_at'] = datetime.now().isoformat()
            ncr['closed_by'] = session.get('full_name')

            DataManager.update_record('ncr_reports', ncr_id, ncr)

            flash('✅ Предписание закрыто', 'success')
            return redirect(url_for('edit_ncr', ncr_id=ncr_id))
        # ─── ОСНОВНОЕ ОБНОВЛЕНИЕ ───
        updated_ncr = ncr.copy()

        def update_if_filled(field):
            value = request.form.get(field)
            if value not in (None, ''):
                updated_ncr[field] = value

        # поля NCR
        fields = [
            'Project', 'Discipline', 'Contractor',
            'technical_supervisor_company',
            'major_object', 'Draw_number',
            'Location', 'Procedure',
            'NCR_Description', 'NCR_grade',
            'inspection_date', 'Correction_acts',
            'closed_date.plan', 'closed_date.actual',
            'Measures', 'NCR_Status'
        ]

        for field in fields:
            update_if_filled(field)


        # ─── ССЫЛКА НА СКАН ───
        signed_url = request.form.get('signed_scan_url', '').strip()
        if signed_url:
            updated_ncr['signed_scan_url'] = signed_url
            updated_ncr['signed_scan_uploaded_at'] = datetime.now().isoformat()
            updated_ncr['signed_scan_uploaded_by'] = session.get('full_name')

        # ─── ФОТО (БЕЗ ИЗМЕНЕНИЙ) ───
        new_photos = []
        photos = request.files.getlist('photos')

        for i, photo in enumerate(photos):
            if photo and photo.filename:
                ext = os.path.splitext(photo.filename)[1]
                filename = f"ncr_{ncr_id}_{len(existing_photos) + i + 1}{ext}"
                photo.save(os.path.join(UPLOAD_FOLDER, filename))
                new_photos.append(filename)

        updated_ncr['photos'] = existing_photos + new_photos

        DataManager.update_record('ncr_reports', ncr_id, updated_ncr)

        flash(
            '⚠️ NCR необходимо распечатать, подписать подрядчиком и загрузить ссылку на подписанный документ',
            'warning'
        )
        flash('✅ NCR обновлён!', 'success')

        return redirect(url_for('edit_ncr', ncr_id=ncr_id))

    return render_template('edit_ncr.html', ncr=ncr)







@app.route('/delete_ncr_photo/<int:ncr_id>/<filename>', methods=['POST'])
@login_required
def delete_ncr_photo(ncr_id, filename):
    ncr = DataManager.get_record_by_id('ncr_reports', ncr_id)
    if not ncr:
        flash('❌ NCR не найден', 'error')
        return redirect(url_for('ncr_list'))

    photos = ncr.get('photos', [])

    if filename in photos:
        photos.remove(filename)

        file_path = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(file_path):
            os.remove(file_path)

        ncr['photos'] = photos
        DataManager.update_record('ncr_reports', ncr_id, ncr)

        flash('🗑️ Фото удалено', 'success')

    return redirect(url_for('edit_ncr', ncr_id=ncr_id))

@app.route('/edit_civil/<int:inspection_id>', methods=['GET', 'POST'])
@login_required
def edit_civil(inspection_id):
    """Редактирование монтажной инспекции"""
    inspection = DataManager.get_record_by_id('civil_inspections', inspection_id)
    
    if not inspection:
        flash('❌ Инспекция не найдена', 'error')
        return redirect(url_for('civil_list'))
    
    if request.method == 'POST':
        updated_inspection = {
    'project': request.form.get('project', ''),
    'major_object': request.form.get('major_object', ''),
    'work_package': request.form.get('work_package', ''),
    'object': request.form.get('object', ''),
    'section': request.form.get('section', ''),
    'work_name': request.form.get('work_name', ''),
    'work_description': request.form.get('work_description', ''),
    'laboratory_number': request.form.get('laboratory_number', ''),
    'quantity': request.form.get('quantity', ''),
    'marking': request.form.get('marking', ''),
    'RFI_number': request.form.get('RFI_number', ''),
    'picket_from': request.form.get('picket_from', ''),
    'picket_to': request.form.get('picket_to', ''),
    'RFI_status': request.form.get('RFI_status', ''),
    'inspection_date': request.form.get('inspection_date', ''),
    'inspector_name_field': request.form.get('inspector_name_field', ''),
    'decision': request.form.get('decision', ''),
    'deviation_comments': request.form.get('deviation_comments', ''),
    'executive_number': request.form.get('executive_number', ''),

    # 🔴 ВАЖНО
    'inspector_id': inspection.get('inspector_id', session.get('user_id')),
    'inspector_name': inspection.get('inspector_name', session.get('full_name'))
}
        
        if DataManager.update_record('civil_inspections', inspection_id, updated_inspection):
            flash('✅ Монтажная инспекция обновлена!', 'success')
            return redirect(url_for('civil_list'))
        else:
            flash('❌ Ошибка при обновлении', 'error')
    
    return render_template('edit_civil.html', inspection=inspection)

@app.route('/edit_daily_report/<int:inspection_id>', methods=['GET', 'POST'])
@login_required
def edit_daily_report(inspection_id):
    """Редактирование монтажной инспекции"""
    inspection = DataManager.get_record_by_id('daily_report', inspection_id)
    
    if not inspection:
        flash('❌ Инспекция не найдена', 'error')
        return redirect(url_for('civil_list'))
    
    if request.method == 'POST':
        updated_inspection = {
            'project': request.form.get('project', ''),
            'major_object': request.form.get('major_object', ''),
            'work_package': request.form.get('work_package', ''),
            'object': request.form.get('object', ''),
            'section': request.form.get('section', ''),
            'work_name': request.form.get('work_name', ''),
            'work_description': request.form.get('work_description', ''),
            'laboratory_number': request.form.get('laboratory_number', ''),
            'quantity': request.form.get('quantity', ''),
            'marking': request.form.get('marking', ''),
            'RFI_number': request.form.get('RFI_number', ''),
            'picket_from': request.form.get('picket_from', ''),
            'picket_to': request.form.get('picket_to', ''),
            'RFI_status': request.form.get('RFI_status', ''),
            'inspection_date': request.form.get('inspection_date', ''),
            'inspector_name_field': request.form.get('inspector_name', ''),
            'decision': request.form.get('decision', ''),
            'deviation_comments': request.form.get('deviation_comments', ''),
            'executive_number': request.form.get('executive_number', '')
        }
        
        if DataManager.update_record('civil_inspections', inspection_id, updated_inspection):
            flash('✅ Монтажная инспекция обновлена!', 'success')
            return redirect(url_for('civil_list'))
        else:
            flash('❌ Ошибка при обновлении', 'error')
    
    return render_template('edit_civil.html', inspection=inspection)

@app.route('/edit_remark/<int:remark_id>', methods=['GET', 'POST'])
@login_required
def edit_remark(remark_id):
    """Редактирование замечания"""
    remark = DataManager.get_record_by_id('remark_inspections', remark_id)
    
    if not remark:
        flash('❌ Замечание не найдено', 'error')
        return redirect(url_for('remark_list'))
    
    if request.method == 'POST':
        updated_remark = {
            'Discipline': request.form.get('Discipline', ''),
            'major_object': request.form.get('major_object', ''),
            'remark_description': request.form.get('remark_description', ''),
            'link_to_normative': request.form.get('link_to_normative', ''),
            'inspection_date': request.form.get(
        'inspection_date',
        remark.get('inspection_date', '')
    ) or remark.get('inspection_date', ''),
            'responsible_person_cont': request.form.get('responsible_person_cont', ''),
            'inspector_name_field': request.form.get('inspector_name_field', ''),
            'remark_closed_date.plan': request.form.get('remark_closed_date.plan', ''),
            'remark_closed_date.actual': request.form.get('remark_closed_date.actual', ''),
            'reason_for_reject': request.form.get('reason_for_reject', ''),
            'Status': request.form.get('Status', ''),
            'inspector_id': remark.get('inspector_id', session.get('user_id')),
    'inspector_name': remark.get('inspector_name', session.get('full_name'))
        }
        
        if DataManager.update_record('remark_inspections', remark_id, updated_remark):
            flash(f'✅ Замечание #{remark_id} обновлено!', 'success')
            return redirect(url_for('remark_list'))
        else:
            flash('❌ Ошибка при обновлении', 'error')
    
    return render_template('edit_remark.html', remark=remark)

@app.route('/delete_remark/<int:remark_id>', methods=['POST'])
@login_required
def delete_remark(remark_id):
    """Удаление замечания"""
    remark = DataManager.get_record_by_id('remark_inspections', remark_id)
    
    if not remark:
        flash('❌ Замечание не найдено', 'error')
        return redirect(url_for('remark_list'))
    
    if DataManager.delete_record('remark_inspections', remark_id):
        flash(f'🗑️ Замечание #{remark_id} удалено!', 'success')
    else:
        flash('❌ Ошибка при удалении', 'error')
    
    return redirect(url_for('remark_list'))

@app.route('/delete_civil/<int:inspection_id>', methods=['POST'])
@login_required
def delete_civil(inspection_id):
    """Удаление монтажной инспекции"""
    inspection = DataManager.get_record_by_id('civil_inspections', inspection_id)
    
    if not inspection:
        flash('❌ Инспекция не найдена', 'error')
        return redirect(url_for('civil_list'))
    
    if DataManager.delete_record('civil_inspections', inspection_id):
        flash(f'🗑️ Монтажная инспекция #{inspection_id} удалена!', 'success')
    else:
        flash('❌ Ошибка при удалении', 'error')
    
    return redirect(url_for('civil_list'))

@app.route('/delete_geodetic/<int:inspection_id>', methods=['POST'])
@login_required
def delete_geodetic(inspection_id):
    """Удаление геодезической инспекции"""
    inspection = DataManager.get_record_by_id('geodetic_inspections', inspection_id)
    
    if not inspection:
        flash('❌ Инспекция не найдена', 'error')
        return redirect(url_for('geodetic_list'))
    
    if DataManager.delete_record('geodetic_inspections', inspection_id):
        flash(f'🗑️ Геодезическая инспекция #{inspection_id} удалена!', 'success')
    else:
        flash('❌ Ошибка при удалении', 'error')
    
    return redirect(url_for('geodetic_list'))

@app.route('/delete_daily_report/<int:report_id>', methods=['POST'])
@login_required
def delete_daily_report(report_id):
    """Удаление ежедневного отчета"""
    report = DataManager.get_record_by_id('daily_reports', report_id)
    
    if not report:
        flash('❌ Отчет не найден', 'error')
        return redirect(url_for('daily_report_list'))
    
    if DataManager.delete_record('daily_reports', report_id):
        flash(f'🗑️ Ежедневный отчет #{report_id} удален!', 'success')
    else:
        flash('❌ Ошибка при удалении', 'error')
    
    return redirect(url_for('daily_report_list'))

@app.route('/delete_ncr/<int:ncr_id>', methods=['POST'])
@login_required
def delete_ncr(ncr_id):
    """Удаление NCR"""
    ncr = DataManager.get_record_by_id('ncr_reports', ncr_id)
    
    if not ncr:
        flash('❌ Отчет не найден', 'error')
        return redirect(url_for('ncr_list'))
    
    if DataManager.delete_record('ncr_reports', ncr_id):
        
        ncr_number = ncr.get('NCR_Number', f'#{ncr_id}')
        flash(f'🗑️ NCR {ncr_number} удален!', 'success')
    else:
        flash('❌ Ошибка при удалении', 'error')
    
    return redirect(url_for('ncr_list'))

@app.route('/')
@login_required
def dashboard():
    """Главная панель"""
    geodetic = DataManager.load_data('geodetic_inspections')
    civil = DataManager.load_data('civil_inspections')
    ncr_reports = DataManager.load_data('ncr_reports') 
    remark_inspections = DataManager.load_data('remark_inspections')
    daily_reports = DataManager.load_data('daily_reports')  # Добавлено

    open_ncr = len([r for r in ncr_reports if r.get('NCR_status') == 'Открыт'])
    closed_ncr = len([r for r in ncr_reports if r.get('NCR_status') == 'Закрыт'])
    
    return render_template('dashboard.html',
                         total_geodetic=len(geodetic),
                         total_civil=len(civil),
                         total_ncr=len(ncr_reports),
                         total_remark=len(remark_inspections),
                         total_daily=len(daily_reports),  # Добавлено
                         open_ncr=open_ncr,
                         closed_ncr=closed_ncr,
                         current_user=session.get('username'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Страница входа"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = DataManager.find_user(username, password)
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['full_name'] = user['full_name']
            flash(f'✅ Добро пожаловать, {user["full_name"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('❌ Неверный логин или пароль', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Выход"""
    full_name = session.get('full_name', 'Пользователь')
    session.clear()
    flash(f'👋 До свидания, {full_name}!', 'success')
    return redirect(url_for('login'))

@app.route('/choose_type')
@login_required
def choose_type():
    """Выбор типа инспекции"""
    return render_template('choose_type.html')

@app.route('/geodetic_list')
@login_required
def geodetic_list():
    """Список геодезических инспекций"""
    inspections = DataManager.load_data('geodetic_inspections')
    
   
    status_filter = request.args.get('status', '')
    decision_filter = request.args.get('decision', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
   
    if status_filter:
        inspections = [i for i in inspections if i.get('status') == status_filter]
    
    if decision_filter:
        inspections = [i for i in inspections if i.get('decision') == decision_filter]
    
    if date_from:
        inspections = [i for i in inspections if i.get('inspection_date', '') >= date_from]
    
    if date_to:
        inspections = [i for i in inspections if i.get('inspection_date', '') <= date_to]
    
    inspections.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return render_template('geodetic_list.html', 
                         inspections=inspections,
                         current_status=status_filter,
                         current_decision=decision_filter,
                         current_date_from=date_from,
                         current_date_to=date_to)

@app.route('/civil_list')
@login_required
def civil_list():
    """Список монтажных инспекций"""
    inspections = DataManager.load_data('civil_inspections')
    
    # Фильтрация
    status_filter = request.args.get('status', '')
    decision_filter = request.args.get('decision', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        inspections = [i for i in inspections if i.get('status') == status_filter]
    if decision_filter:
        inspections = [i for i in inspections if i.get('decision') == decision_filter]
    if date_from:
        inspections = [i for i in inspections if i.get('inspection_date', '') >= date_from]
    if date_to:
        inspections = [i for i in inspections if i.get('inspection_date', '') <= date_to]
    
    inspections.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return render_template('civil_list.html', 
                         inspections=inspections,
                         current_status=status_filter,
                         current_decision=decision_filter,
                         current_date_from=date_from,
                         current_date_to=date_to)

@app.route('/ncr_list')
@login_required
def ncr_list():
    """Список отчетов о несоответствии"""
    ncr_reports = DataManager.load_data('ncr_reports')
    
    # Фильтрация
    status_filter = request.args.get('NCR_status', '')
    priority_filter = request.args.get('priority', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        ncr_reports = [r for r in ncr_reports if r.get('NCR_status') == status_filter]
    if priority_filter:
        ncr_reports = [r for r in ncr_reports if r.get('priority') == priority_filter]
    if date_from:
        ncr_reports = [r for r in ncr_reports if r.get('inspection_date', '') >= date_from]
    if date_to:
        ncr_reports = [r for r in ncr_reports if r.get('inspection_date', '') <= date_to]
    
    ncr_reports.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return render_template('ncr_list.html', 
                         ncr_reports=ncr_reports,
                         current_status=status_filter,
                         current_priority=priority_filter,
                         current_date_from=date_from,
                         current_date_to=date_to)

@app.route('/print_ncr/<int:ncr_id>')
@login_required
def print_ncr(ncr_id):
    ncr = DataManager.get_record_by_id('ncr_reports', ncr_id)

    return render_template(
        'print_ncr.html',
        ncr=ncr,
        total_pages=3  # ← пока вручную
    )


@app.route('/remark_list')
@login_required
def remark_list():
    """Список отчетов о замечаниях"""
    remark_reports = DataManager.load_data('remark_inspections')
    
    # Фильтрация
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    if status_filter:
        remark_reports = [r for r in remark_reports if r.get('Status') == status_filter]
    if date_from:
        remark_reports = [r for r in remark_reports if r.get('inspection_date', '') >= date_from]
    if date_to:
        remark_reports = [r for r in remark_reports if r.get('inspection_date', '') <= date_to]
    
    remark_reports.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return render_template('remark_list.html', 
                         remark_reports=remark_reports,
                         current_status=status_filter,
                         current_date_from=date_from,
                         current_date_to=date_to)

@app.route('/daily_report_list')
@login_required
def daily_report_list():
    """Список ежедневных отчетов"""
    reports = DataManager.load_data('daily_reports')
    
    # Фильтрация
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    author_filter = request.args.get('author', '')
    
    if date_from:
        reports = [r for r in reports if r.get('report_date', '') >= date_from]
    if date_to:
        reports = [r for r in reports if r.get('report_date', '') <= date_to]
    if author_filter:
        reports = [r for r in reports if author_filter.lower() in r.get('author', '').lower()]
    
    reports.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return render_template('daily_report_list.html', 
                         reports=reports,
                         current_date_from=date_from,
                         current_date_to=date_to,
                         current_author=author_filter)

@app.route('/view_daily_report/<int:report_id>')
@login_required
def view_daily_report(report_id):
    """Просмотр ежедневного отчета"""
    report = DataManager.get_record_by_id('daily_reports', report_id)
    if not report:
        flash('❌ Отчет не найден', 'error')
        return redirect(url_for('daily_report_list'))
    
    return render_template('view_daily_report.html', report=report)

@app.route('/new_geodetic', methods=['GET', 'POST'])
@login_required
def new_geodetic():
    """Создание геодезической инспекции"""
    if request.method == 'POST':
        inspection = {
            'project': request.form.get('project', ''),
            'major_object': request.form.get('major_object', ''),
            'work_package': request.form.get('work_package', ''),
            'inspection_date': request.form.get('inspection_date', ''),
            'object': request.form.get('object', ''),
            'section': request.form.get('section', ''),
            'work_name': request.form.get('work_name', ''),
            'quantity_project': request.form.get('quantity_project', ''),
            'quantity_actual': request.form.get('quantity_actual', ''),
            'picket_from': request.form.get('picket_from', ''),
            'picket_to': request.form.get('picket_to', ''),
            'inspector_name_field': request.form.get('inspector_name_field', ''),
            'decision': request.form.get('decision', ''),
            'deviation_comments': request.form.get('deviation_comments', ''),
            'inspector_id': session.get('user_id'),
            'inspector_name': session.get('full_name')
        }
        
        if DataManager.add_record('geodetic_inspections', inspection):
            flash('✅ Геодезическая инспекция создана!', 'success')
            return redirect(url_for('geodetic_list'))
        else:
            flash('❌ Ошибка при создании', 'error')
    
    return render_template('new_geodetic.html')

@app.route('/new_civil', methods=['GET', 'POST'])
@login_required
def new_civil():
    """Создание инспекции по монтажным работам"""
    if request.method == 'POST':
        inspection = {
            'project': request.form.get('project', ''),
            'major_object': request.form.get('major_object', ''),
            'work_package': request.form.get('work_package', ''),
            'object': request.form.get('object', ''),
            'section': request.form.get('section', ''),
            'work_name': request.form.get('work_name', ''),
            'work_description': request.form.get('work_description', ''),
            'laboratory_number': request.form.get('laboratory_number', ''),
            'quantity': request.form.get('quantity', ''),
            'marking': request.form.get('marking', ''),
            'RFI_number': request.form.get('RFI_number', ''),
            'picket_from': request.form.get('picket_from', ''),
            'picket_to': request.form.get('picket_to', ''),
            'RFI_status': request.form.get('RFI_status', ''),
            'inspection_date': request.form.get('inspection_date', ''),
            'inspector_name_field': request.form.get('inspector_name_field', ''),
            'decision': request.form.get('decision', ''),
            'deviation_comments': request.form.get('deviation_comments', ''),
            'executive_number': request.form.get('executive_number', ''),
            'inspector_id': session.get('user_id'),
            'inspector_name': session.get('full_name'),
        }

        
        if DataManager.add_record('civil_inspections', inspection):
            flash('✅ Civil inpection создана!', 'success')
            return redirect(url_for('civil_list'))
        else:
            flash('❌ Ошибка при создании', 'error')
    
    return render_template('new_civil.html')

@app.route('/new_remark', methods=['GET', 'POST'])
@login_required
def new_remark():
    """Создание отчета о замечании"""
    if request.method == 'POST':
        remark_report = {
            'Discipline': request.form.get('Discipline', ''),
            'major_object': request.form.get('major_object', ''),
            'remark_description': request.form.get('remark_description', ''),
            'link_to_normative': request.form.get('link_to_normative', ''),
            'inspection_date': request.form.get('inspection_date', ''),
            'responsible_person_cont': request.form.get('responsible_person_cont', ''),
            'inspector_name_field': request.form.get('inspector_name_field', ''),
            'remark_closed_date.plan': request.form.get('remark_closed_date.plan', ''),
            'remark_closed_date.actual': request.form.get('remark_closed_date.actual', ''),
            'reason_for_reject': request.form.get('reason_for_reject', ''),
            'Status': request.form.get('Status', ''),
            'inspector_id': session.get('user_id'),
            'inspector_name': session.get('full_name')
        }
        
        if DataManager.add_record('remark_inspections', remark_report):
            flash('✅ Отчет о замечании создан!', 'success')
            return redirect(url_for('remark_list'))
        else:
            flash('❌ Ошибка при создании отчета', 'error')
    
    return render_template('new_remark.html')

@app.route('/new_ncr', methods=['GET', 'POST'])
@login_required
def new_ncr():
    try:
        generated_ncr_number = generate_ncr_number()
    except Exception as e:
        print("Ошибка генерации NCR:", e)
        flash("Ошибка генерации номера NCR", "error")
        return redirect(url_for('dashboard'))
    def generate_ncr_number():
    prefix = "TSFS-AGMK-NCR-"
    try:
        ncr_list = DataManager.load_data('ncr_reports')
    except Exception as e:
        print("Ошибка загрузки NCR:", e)
        return prefix + "0001"

    numbers = []

    for ncr in ncr_list:
        num = ncr.get('NCR_Number', '')
        if num.startswith(prefix):
            try:
                numbers.append(int(num.replace(prefix, '')))
            except ValueError:
                continue  # пропускаем битые номера

    next_number = max(numbers) + 1 if numbers else 1
    return prefix + f"{next_number:04d}"
    

    if request.method == 'POST':

        ncr_report = {
            'NCR_Number': generated_ncr_number,
            'Project': request.form.get('Project', ''),
            'Contractor': request.form.get('Contractor', ''),
            'technical_supervisor_company': request.form.get('technical_supervisor_company', ''),
            'Discipline': request.form.get('Discipline', ''),
            'major_object': request.form.get('major_object', ''),
            'Draw_number': request.form.get('Draw_number', ''),
            'Location': request.form.get('Location', ''),
            'Procedure': request.form.get('Procedure', ''),
            'NCR_Description': request.form.get('NCR_Description', ''),
            'NCR_grade': request.form.get('NCR_grade', ''),
            'inspection_date': request.form.get('inspection_date', ''),
            'Correction_acts': request.form.get('Correction_acts', ''),
            'closed_date.plan': request.form.get('closed_date.plan', ''),
            'closed_date.actual': request.form.get('closed_date.actual', ''),
            'Measures': request.form.get('Measures', ''),
            'inspector_name_field': request.form.get('inspector_name_field', ''),
            'NCR_Status': request.form.get('NCR_Status', ''),
            'inspector_id': session.get('user_id'),
            'inspector_name': session.get('full_name'),
            'photos': []
        }

        if not DataManager.add_record('ncr_reports', ncr_report):
            flash('❌ Ошибка при создании отчета', 'error')
            return redirect(url_for('new_ncr'))
        flash(
    '⚠️ Внимание: NCR необходимо распечатать, подписать подрядчиком, '
    'отсканировать (вложить по адресу: PMO-AGMK\Tailing dams (XX) - Документы\. TSF reports\Файлы по ТН\Предписания\Сканы предписания) , сделать рассылку после подписания.',
    'warning'
)

        # фото — без изменений
        all_ncr = DataManager.load_data('ncr_reports')
        ncr_id = max(all_ncr, key=lambda x: x['id'])['id']

        photos = request.files.getlist('photos')
        photo_filenames = []

        for i, photo in enumerate(photos):
            if photo and photo.filename:
                filename = secure_filename(photo.filename)
                ext = os.path.splitext(filename)[1]
                saved_name = f"ncr_{ncr_id}_{i+1}{ext}"
                photo.save(os.path.join(UPLOAD_FOLDER, saved_name))
                photo_filenames.append(saved_name)

        all_ncr[-1]['photos'] = photo_filenames
        DataManager.save_data('ncr_reports', all_ncr)

        flash(f'✅ Отчет {generated_ncr_number} создан!', 'success')
        return redirect(url_for('ncr_list'))
    def get_last_contractor():
        ncr_list = DataManager.load_data('ncr_reports')
        if not ncr_list:
            return ''
        return ncr_list[-1].get('Contractor', '')
    def get_last_technical_supervisor_company():
        ncr_list = DataManager.load_data('ncr_reports')
        if not ncr_list:
            return ''
        return ncr_list[-1].get('technical_supervisor_company', '')
    

    return render_template(
    'new_ncr.html',
    generated_ncr_number=generate_ncr_number(),
    generated_contractor=get_last_contractor(),
    generated_technical_supervisor_company=get_last_technical_supervisor_company()
)
   



@app.route('/daily_report', methods=['GET', 'POST'])
@login_required
def daily_report():
    """Создание ежедневного отчета"""
    if request.method == 'POST':
        # Обработка фотографий (без изменений)
        photos_data = []
        photo_captions = request.form.getlist('photo_captions[]')
        
        for key, value in request.form.items():
            if key.startswith('photo_data_'):
                index = int(key.split('_')[-1])
                caption = photo_captions[index] if index < len(photo_captions) else f'Фото {index + 1}'
                photos_data.append({
                    'data': value,
                    'caption': caption
                })
        
        daily_report = {
            # Основная информация
            'report_date': request.form.get('report_date', ''),
            'project_name': request.form.get('project_name', ''),
            'location': request.form.get('location', ''),
            'author': request.form.get('author', ''),
            'shift': request.form.get('shift', ''),
            'weather': request.form.get('weather', ''),
            
            # НОВЫЕ ПОЛЯ
            'personnel': request.form.get('personnel', ''),
            'material_placement': request.form.get('material_placement', ''),
            
            # Техника (без изменений)
            'equipment_data': {
                'names': request.form.getlist('equipment_name[]'),
                'damba': request.form.getlist('damba[]'),
                'vodovod': request.form.getlist('vodovod[]'),
                'gpp': request.form.getlist('gpp[]'),
                'pulpovod': request.form.getlist('pulpovod[]'),
                'raspred': request.form.getlist('raspred[]'),
                'total': request.form.getlist('total[]'),
                'total_site': request.form.getlist('total_site[]')
            },
            
            # ИСПРАВЛЕННЫЕ ИМЕНА: Выполняемые работы
            'works_data': {
                'areas': request.form.getlist('work_area[]'),
                'ch_from': request.form.getlist('work_ch_from[]'),
                'ch_to': request.form.getlist('work_ch_to[]'),
                'work_types': request.form.getlist('work_type[]'),
                'work_descriptions': request.form.getlist('work_description[]')
            },
            
            # НОВАЯ СТРУКТУРА: Поставка материалов
            'materials_data': {
                'areas': request.form.getlist('material_area[]'),
                'material_types': request.form.getlist('material_type[]'),
                'descriptions': request.form.getlist('material_description[]')
            },
            
            # НОВАЯ СТРУКТУРА: Контроль качества
            'quality_data': {
                'areas': request.form.getlist('qc_area[]'),
                'ch_from': request.form.getlist('qc_ch_from[]'),
                'ch_to': request.form.getlist('qc_ch_to[]'),
                'work_types': request.form.getlist('qc_work_type[]'),
                'descriptions': request.form.getlist('qc_description[]')
            },
            
            'remarks': request.form.get('remarks', ''),
            'photos_data': photos_data,
            'inspector_id': session.get('user_id'),
            'inspector_name': session.get('full_name')
        }
        
        if DataManager.add_record('daily_reports', daily_report):
            flash('✅ Ежедневный отчет создан!', 'success')
            return redirect(url_for('daily_report_list'))
        else:
            flash('❌ Ошибка при создании отчета', 'error')
    
    return render_template('daily_report.html')

@app.route('/api/weather')
@login_required
def api_weather():
    """API для получения текущей погоды"""
    weather = get_weather_data()
    return {'weather': weather, 'timestamp': datetime.now().isoformat()}

@app.route('/debug_data')
@login_required
def debug_data():
    return "ok"

@app.route("/healthz")
def healthz():
    return "ok", 200

def create_test_users():
    """Создание тестовых пользователей"""
    users = DataManager.load_data('users')
    if not users:
        test_users = [
            {
                'id': 1,
                'username': 'madiyar',
                'password': '123456',
                'full_name': 'Мадияр Адильжанов',
                'position': 'Старший геодезист',
                'created_at': datetime.now().isoformat()
            },
            {
                'id': 2,
                'username': 'said777',
                'password': '123456',
                'full_name': 'Said Djurabekov',
                'position': 'Геодезист-инспектор',
                'created_at': datetime.now().isoformat()
            }
        ]
        DataManager.save_data('users', test_users)
        print("✅ Созданы тестовые пользователи")







if __name__ == '__main__':  
    app.run(debug=True)





