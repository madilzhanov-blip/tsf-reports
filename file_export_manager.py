# -*- coding: utf-8 -*-
import os
import shutil
import smtplib
import zipfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import datetime
import io

class FileExportManager:
    def __init__(self, export_folder="exports", network_folder=None):
        """
        export_folder: локальная папка для сохранения
        network_folder: сетевая папка (например: r"\\\\server\\shared\\reports")
        """
        self.export_folder = export_folder
        self.network_folder = network_folder
        self.create_folders()
    
    def create_folders(self):
        """Создание необходимых папок"""
        folders = [
            self.export_folder,
            os.path.join(self.export_folder, "civil"),
            os.path.join(self.export_folder, "geodetic"), 
            os.path.join(self.export_folder, "ncr"),
            os.path.join(self.export_folder, "remarks"),
            os.path.join(self.export_folder, "archive"),
            os.path.join(self.export_folder, "test")
        ]
        
        for folder in folders:
            if not os.path.exists(folder):
                os.makedirs(folder)
                print(f"✅ Создана папка: {folder}")
    
    def save_excel_file(self, workbook, inspection_type, inspection_id=None):
        """Сохранение Excel файла локально"""
        try:
            # Генерируем имя файла
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if inspection_id:
                filename = f"{inspection_type}_inspection_{inspection_id}_{timestamp}.xlsx"
            else:
                filename = f"{inspection_type}_report_{timestamp}.xlsx"
            
            # Путь к подпапке
            subfolder = inspection_type.lower()
            subfolder_path = os.path.join(self.export_folder, subfolder)
            
            # Создаем подпапку если её нет
            if not os.path.exists(subfolder_path):
                os.makedirs(subfolder_path)
                print(f"✅ Создана подпапка: {subfolder_path}")
            
            # Полный путь к файлу
            filepath = os.path.join(subfolder_path, filename)
            
            # Сохраняем файл
            workbook.save(filepath)
            
            print(f"✅ Файл сохранен: {filepath}")
            return True, filepath, filename
            
        except Exception as e:
            print(f"❌ Ошибка сохранения: {e}")
            return False, None, None
    
    def copy_to_network_folder(self, filepath, filename):
        """Копирование в сетевую папку"""
        if not self.network_folder:
            return False, "Сетевая папка не настроена"
        
        try:
            if not os.path.exists(self.network_folder):
                return False, f"Сетевая папка недоступна: {self.network_folder}"
            
            network_path = os.path.join(self.network_folder, filename)
            shutil.copy2(filepath, network_path)
            
            print(f"✅ Файл скопирован в сеть: {network_path}")
            return True, f"Файл скопирован в: {network_path}"
            
        except Exception as e:
            print(f"❌ Ошибка копирования в сеть: {e}")
            return False, f"Ошибка: {e}"
    
    def create_archive(self, inspection_type, start_date=None, end_date=None):
        """Создание архива отчетов за период"""
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            archive_name = f"{inspection_type}_archive_{timestamp}.zip"
            archive_path = os.path.join(self.export_folder, "archive", archive_name)
            
            source_folder = os.path.join(self.export_folder, inspection_type.lower())
            
            with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(source_folder):
                    for file in files:
                        if file.endswith('.xlsx'):
                            # Фильтрация по датам (если нужно)
                            file_path = os.path.join(root, file)
                            arcname = os.path.relpath(file_path, source_folder)
                            zipf.write(file_path, arcname)
            
            print(f"✅ Архив создан: {archive_path}")
            return True, archive_path
            
        except Exception as e:
            print(f"❌ Ошибка создания архива: {e}")
            return False, None

class EmailSender:
    def __init__(self, smtp_server, smtp_port, username, password):
        """
        Настройки для отправки email
        smtp_server: например "smtp.gmail.com" или "smtp.company.com"
        smtp_port: обычно 587 для TLS или 465 для SSL
        """
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.username = username
        self.password = password
    
    def send_report(self, to_emails, subject, body, attachment_path=None):
        """Отправка отчета по email"""
        try:
            # Создаем сообщение
            msg = MIMEMultipart()
            msg['From'] = self.username
            msg['To'] = ", ".join(to_emails) if isinstance(to_emails, list) else to_emails
            msg['Subject'] = subject
            
            # Добавляем текст
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            # Добавляем вложение
            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                filename = os.path.basename(attachment_path)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {filename}'
                )
                msg.attach(part)
            
            # Отправляем
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.username, self.password)
            server.send_message(msg)
            server.quit()
            
            print(f"✅ Email отправлен: {to_emails}")
            return True, "Email отправлен успешно"
            
        except Exception as e:
            print(f"❌ Ошибка отправки email: {e}")
            return False, f"Ошибка: {e}"

# Конфигурация для разных компаний
EMAIL_CONFIGS = {
    "gmail": {
        "smtp_server": "smtp.gmail.com",
        "smtp_port": 587
    },
    "outlook": {
        "smtp_server": "smtp-mail.outlook.com", 
        "smtp_port": 587
    },
    "corporate": {
        "smtp_server": "mail.company.com",  # Замени на корпоративный сервер
        "smtp_port": 587
    }
}

def test_export_system():
    """Тестирование системы экспорта"""
    print("🧪 Тестирование системы экспорта...")
    
    # Создаем менеджер экспорта
    export_manager = FileExportManager()
    
    # Тестируем создание папок
    print("✅ Папки созданы")
    
    # Тестируем с фиктивными данными
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Тестовый отчет"
        ws['A2'] = "Дата создания"
        ws['B2'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        success, filepath, filename = export_manager.save_excel_file(wb, "test", 999)
        
        if success:
            print(f"✅ Тестовый файл создан: {filename}")
        
    except ImportError:
        print("⚠️ openpyxl не найден, но структура папок готова")
    
    print("🎉 Тест завершен!")

if __name__ == "__main__":
    test_export_system()